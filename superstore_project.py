import pandas as pd
import numpy as np
import matplotlib.pyplot as mp
import seaborn as sb
from openpyxl import load_workbook
from openpyxl.styles import  Font, PatternFill


# load the file

rawfile = pd.read_excel("C:\\Users\\golla\\OneDrive\\Desktop\Python Data analyst Project\\Sample - Superstore.xlsx" )

df = pd.DataFrame(rawfile)
#print(df)

#Extract only top 5 rows
#print(df.head())

# perform the statistiocal operatons using "Describe function"
#print(df.describe())

print(df.info())
# Sheet:1----product Wise Sale
ProductWiseID_Sale = df.groupby(["Product ID","Category"])[["Quantity","Sales","Profit"]].sum().round(2).reset_index().sort_values(by="Category",ascending=True)
df["Order Date"] = pd.to_datetime(df["Order Date"],errors= "coerce")
highest_Date = df["Order Date"].dt.day.max()
print(highest_Date)
ProductWiseID_Sale["Avg Sale per day"] = ProductWiseID_Sale["Sales"]/highest_Date
ProductWiseID_Sale["Estimated Sale EOM"]= (ProductWiseID_Sale["Avg Sale per day"]*31).round(2)
Cat_Total = []
subtotal = []
for cat, group in ProductWiseID_Sale.groupby("Category"):
    Cat_Total.append(group.copy())
    sales = group["Sales"].sum()
    Esm_Sale = group["Avg Sale per day"].sum()*31
    profit = group["Profit"].sum()
    Cat_Total.append(group)
    Row_total={
        "Product ID" : f"{cat} total",
        "Category" : "",
        "Quantity" : group["Quantity"].sum(),
        "Sales" : sales,
         "Profit" : profit,
        "Avg Sale per day" : sales/highest_Date,
        "Estimated Sale EOM" : Esm_Sale,
       
    }
    subtotal.append(pd.DataFrame([Row_total]))
# Grand total row calculation 
Grand_Total = []
total_sales = ProductWiseID_Sale["Sales"].sum()
total_EME_sale = ProductWiseID_Sale["Estimated Sale EOM"].sum()
grand_total = {
    "Product ID" :"Grand Total",
    "Category" : "",
    "Quantity" : ProductWiseID_Sale["Quantity"].sum(),
    "Sales" : ProductWiseID_Sale["Sales"].sum(),
    "Profit" : ProductWiseID_Sale["Profit"].sum(),
    "Avg Sale per day" : ProductWiseID_Sale["Sales"].sum()/highest_Date,
    "Estimated Sale EOM" : ProductWiseID_Sale["Avg Sale per day"].sum()*31,
    "Profit" : ProductWiseID_Sale["Profit"].sum()
}
Grand_Total.append(pd.DataFrame([grand_total]))
Final_Results = pd.concat([ProductWiseID_Sale]+ subtotal+ Grand_Total ,ignore_index=True)
#-------------------------------------------------------------------------------------------------------
# Sheet:2--Category Wise Sales
Category_Wise_Sale = df.groupby(["Category","Sub-Category"])[["Quantity","Sales","Profit"]].sum().round(2).reset_index().sort_values(by="Category", ascending=True)

Category_Grand_total =[]
Cat_grand_total = {
    "Category":"Grand Total",
    "Sub-Category": "",
    "Quantity" : Category_Wise_Sale["Quantity"].sum(),
    "Sales" : Category_Wise_Sale["Sales"].sum(),
    "Profit" : Category_Wise_Sale["Profit"].sum()
}
Category_Grand_total.append(pd.DataFrame([Cat_grand_total]))
Category_Wise_Sale = pd.concat([Category_Wise_Sale]+Category_Grand_total,ignore_index=True)
#---------------------------------------------------------------------------------------------------------
#Sheet:3-- Customer Wise sale Dashboard
Customer_Wise_Sale = df.groupby(["Customer ID","Customer Name"])[["Quantity","Sales","Profit"]].sum().round(2).reset_index().sort_values(by="Customer Name",ascending=True)
Customer_Grand_total =[]
Cust_grand_total = {
    "Customer ID":"Grand Total",
    "Customer Name": "",
    "Quantity" : Customer_Wise_Sale["Quantity"].sum(),
    "Sales" : Customer_Wise_Sale["Sales"].sum(),
    "Profit" : Customer_Wise_Sale["Profit"].sum()
}
Customer_Grand_total.append(pd.DataFrame([Cust_grand_total]))
Customer_Wise_Sale = pd.concat([Customer_Wise_Sale]+Customer_Grand_total,ignore_index=True)
#-----------------------------------------------------------------------------------------------------------------
# Sheet:4--State Wise Sales Dashboard
df["State Name"] = df["State"]
State_wise_Sales = df.groupby(["Country", "Region","State Name"])[["Quantity","Sales","Profit"]].sum().round(2).reset_index().sort_values(by="State Name",ascending=True)
State_Grand_total =[]
State_grand_total = {
    "Country":"Grand Total",
    "Region": "",
    "State Name" : "",
    "Quantity" : State_wise_Sales["Quantity"].sum(),
    "Sales" : State_wise_Sales["Sales"].sum(),
    "Profit" : State_wise_Sales["Profit"].sum()
}
State_Grand_total.append(pd.DataFrame([State_grand_total]))
State_wise_Sales = pd.concat([State_wise_Sales]+State_Grand_total,ignore_index=True)

# Highleting total and Grand total rows
def highlets_row(sheetname):
    def total_row_highlets(row):
        value = str(row.get(sheetname,""))
        if "Grand" in value:
             return ["font-weight: bold; background-color: #4f81bd"] * len(row)
        elif "total" in value:
            return ["font-weight: bold; background-color: #f9e79f"] *len(row)
        else:
            return [""]*len(row)
    return total_row_highlets
Final_Results = Final_Results.style.apply(highlets_row("Product ID"),axis=1)
Category_Wise_Sale = Category_Wise_Sale.style.apply(highlets_row("Category"),axis=1)
Customer_Wise_Sale = Customer_Wise_Sale.style.apply(highlets_row("Customer ID"),axis=1)
State_wise_Sales = State_wise_Sales.style.apply(highlets_row("Country"),axis=1)

outputfile = r"C:\\Users\\golla\\OneDrive\\Desktop\\Python Data analyst Project\\Retail_Project_Output.xlsx"
with pd.ExcelWriter(outputfile, engine="openpyxl") as writer:
    Final_Results.to_excel(writer, sheet_name="PRODUCT ID", index=False)
    Category_Wise_Sale.to_excel(writer, sheet_name="Category Wise",index=False)
    Customer_Wise_Sale.to_excel(writer,sheet_name="Customer Wise",index=False)
    State_wise_Sales.to_excel(writer, sheet_name="State Wise",index=False)
wb = load_workbook(outputfile)
for sheet in wb.sheetnames:
    ws= wb[sheet]
    for col in ws.columns:
        max_len = 0
        colum = col[0].column_letter
        for cell in col:
            try:
                 if cell.valuex:
                    max_len = max(max_len, len(str(cell.values)))
            except:
                pass
        adjust_len = max_len+20
        ws.column_dimensions[colum].width = adjust_len
wb.save(outputfile)
wb = load_workbook(outputfile)
header_colors = {
    "Product ID": "4F81BD",    # Blue
    "Category": "4F81BD",      # Blue
    "Sales": "92D050",         # Green
    "Quantity": "92D050",       # Green
    "Avg Sale per day": "00B0F0", # Sky Blue
    "Estimated EOM": "00B0F0",  # Sky
    "Sales Gap" :"C9198E",        # Purple
    "Profit" : "92D050"
}
for sheet in wb.sheetnames:
    ws = wb[sheet]
    for cell in ws[1]:
        col_name = cell.value
        if col_name in header_colors:
            color = header_colors[col_name]
            cell.fill = PatternFill(start_color= color, end_color=color ,fill_type="darkDown")
            cell.font = Font(color="000000", bold=True)
        else:
            cell.fill = PatternFill(start_color="4400ff", end_color="4400ff", fill_type="solid" )
            cell.font = Font(color="000000",bold = True)

wb.save(outputfile)